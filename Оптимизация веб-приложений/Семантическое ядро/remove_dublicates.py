import sys
from openpyxl import load_workbook


keywords_list = []

wb = load_workbook(sys.argv[1])
sheet = wb[wb.sheetnames[0]]

i=2
cell = sheet["D"+str(i)]
keywords_list.append(cell.value)
while cell.value!=None:
	cell = sheet["D"+str(i)]
	keywords_list.append(cell.value)
	print(cell.value)
	i+=1

keywords_list.remove(None)
wo_dubles = sorted(set(keywords_list))
k=2
for value in wo_dubles:
	sheet["D"+str(k)].value = value
	k+=1

while k<=i:
	sheet["D"+str(k)].value = ""
	k+=1

wb.save(sys.argv[1])