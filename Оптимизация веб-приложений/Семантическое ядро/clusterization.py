import sys
from openpyxl import load_workbook


keywords_list = []

wb = load_workbook(sys.argv[1])
sheet = wb[wb.sheetnames[0]]

i=3
cell = sheet[sys.argv[2]+str(i)]
keywords_list.append(cell.value)

while cell.value!=None:
	cell = sheet[sys.argv[2]+str(i)]
	keywords_list.append(cell.value)
	print(cell.value)
	i+=1

keywords_list.remove(None)
themes = sheet[sys.argv[2]+"2"].value.split('/')
themed = []
for phrase in keywords_list:
	for theme in themes:
		if phrase.lower().find(theme.lower(), 0)!=-1:
			themed.append(phrase)
themed.sort()
k=3
for value in themed:
	sheet[sys.argv[2]+str(k)].value = value
	k+=1

while k<=i:
	sheet[sys.argv[2]+str(k)].value = ""
	k+=1

wb.save(sys.argv[1])