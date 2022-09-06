import sys
from openpyxl import load_workbook


keywords_list = []

wb = load_workbook(sys.argv[1])
sheet = wb[wb.sheetnames[0]]

i=2
cell = sheet["E"+str(i)]
keywords_list.append(cell.value)

while cell.value!=None:
	cell = sheet["E"+str(i)]
	keywords_list.append(cell.value)
	print(cell.value)
	i+=1

keywords_list.remove(None)
wo_minus=[]
for phrase in keywords_list:
	if phrase.lower().find(sys.argv[2].lower(), 0)==-1:
		wo_minus.append(phrase)
wo_minus.sort()
k=2
for value in wo_minus:
	sheet["E"+str(k)].value = value
	k+=1

while k<=i:
	sheet["E"+str(k)].value = ""
	k+=1

wb.save(sys.argv[1])