#/usr/bin/python3
import sys
from openpyxl import load_workbook
wb = load_workbook(sys.argv[1])
sheet = wb[wb.sheetnames[0]]
data = open(sys.argv[2], 'r').read().splitlines()

i=142
for line in data:
	print(line)
	print(f'\n{i}\n')
	sheet.cell(row=i, column=3).value = line
	i+=1
wb.save(sys.argv[1])

